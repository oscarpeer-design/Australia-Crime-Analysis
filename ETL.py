import openpyxl
from openpyxl import load_workbook
from Neural_Network_Application import NeuralNetwork

class Extract_EXCEL:
    def __init__(self, spreadsheet_name, main_workbook_name):
        self.spreadsheet_name = spreadsheet_name
        self.main_workbook_name = main_workbook_name
        self.data = None

    def data_is_read(self):
        return self.data is not None

    def open_main_workbook(self):
        try:
            #Load existing workbook
            self.workbook = load_workbook(
                filename=self.main_workbook_name,
                read_only=True,
                data_only=True
            )
        except Exception as e:
            raise RuntimeError(f"Failed to open workbook: {e}")

    def read_sheet_columns(self, requested_headers):
        if not hasattr(self, "workbook"):
            raise RuntimeError("Workbook not opened")

        if self.spreadsheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{self.spreadsheet_name}' not found")

        sheet = self.workbook[self.spreadsheet_name]

        # Read header row
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        header_index = {
            header: idx
            for idx, header in enumerate(header_row)
            if header is not None
        }

        missing = [h for h in requested_headers if h not in header_index]
        if missing:
            raise ValueError(f"Columns not found: {missing}")

        selected_columns = []

        for header in requested_headers:
            col_idx = header_index[header] + 1  # 1-based index
            column_data = [
                row[0] for row in sheet.iter_rows(
                    min_col=col_idx,
                    max_col=col_idx,
                    values_only=True
                )
            ]
            selected_columns.append(column_data)

        self.data = selected_columns

    def get_data(self):
        return self.data

def sort_data(data):
    #print(f"Data: {data}") #Debugging output statement
    #Sorts a list of lists
    for data_list in data:
        data_list.sort()
    return data

def remove_column_headings(data, keep_first_column):
    columns = data
    if keep_first_column is True:
        columns = data[1:]
        
    for i in range(len(columns)):
        column = columns[i]
        column = column[1:]
        columns[i] = column 

    if keep_first_column is True:
        data = [data[0]] 
        for column in columns :
            data.append(column)
    return data

def standardise(values):
    mean = sum(values)/len(values)
    std = (sum((x-mean)**2 for x in values)/len(values))**0.5
    return [(x-mean)/std for x in values]

def standardise_values(data):
    for values in data:
        values = standardise(values)
    return data

def UseExtract_Excel(spreadsheet_name, workbook_name, column_names, remove_headings, keep_first_column, sort = True, standardisation = True):
    etl = Extract_EXCEL(spreadsheet_name, workbook_name)
    etl.open_main_workbook()
    etl.read_sheet_columns(column_names)
    if etl.data_is_read():
        if remove_headings is True:
            data = etl.get_data()
            data = remove_column_headings(data, keep_first_column)
        if sort is True:
           data = sort_data(data)
        if standardisation is True:
            data = standardise_values(data[1:])
        return data
    return None

def Load_NeuralNetwork(variable_data, result_data, learning_rate, size_hidden, size_output):
    neural_network = NeuralNetwork(variable_data, result_data, learning_rate, size_hidden, size_output)
    return neural_network